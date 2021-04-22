from openpyxl import Workbook

from govalidator.validators import OntologyValidator
from openpyxl.utils import get_column_letter

from govalidator.gotemplate import Gotemplate


class Gogenerator(Gotemplate):
    def __init__(
        self,
        output,
        **kwargs
    ):
        super(Gogenerator, self).__init__(**kwargs)
        self.output = output

    def generate(self):
        wb = Workbook()
        current_data_column = 1
        current_ontology_column = 1
        current_readme_row = 1
        readme_sheet = wb.active
        readme_sheet.title = "README"
        data_sheet = wb.create_sheet(title="Data")
        ontology_sheet = None
        for column_name, validator in self.validators.items():
            readme_sheet.cell(column=1, row=current_readme_row, value=validator.describe(column_name))
            current_readme_row += 1
            data_sheet.cell(column=current_data_column, row=1, value=column_name)
            if isinstance(validator, OntologyValidator):
                if not ontology_sheet:
                    ontology_sheet = wb.create_sheet(title="Ontologies")
                data_validation = validator.generate(get_column_letter(current_data_column), get_column_letter(current_ontology_column), ontology_sheet)
                current_ontology_column += 1
            else:
                data_validation = validator.generate(get_column_letter(current_data_column))
            if data_validation:
                data_sheet.add_data_validation(data_validation)
            current_data_column += 1
        wb.save(filename=self.output)

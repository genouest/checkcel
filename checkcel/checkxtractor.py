from checkcel import logs
from openpyxl import load_workbook


class Checkxtractor(object):
    """ Extract validation value from xlsx file (only) """
    def __init__(self, source, output, sheet=0, row=0):
        self.logger = logs.logger
        self.source = source
        self.output = output
        self.sheet = int(sheet)
        self.row = int(row)
        self.columns_list = []
        self.validation_list = []
        self.names = {}
        self.wb = None
        self.used_validators = set()

    def extract(self):
        self.wb = load_workbook(self.source)
        for name in self.wb.defined_names.definedName:
            self.names[name.name] = name.destinations
        ws = self.wb.worksheets[self.sheet]
        columns_list = []
        validation_dict = {}
        # Get active columns with name
        for cell in ws[self.row + 1]:
            if cell.value:
                columns_list.append(cell.value)
        if not columns_list:
            raise Exception("TODO")
        for validation in ws.data_validations.dataValidation:
            if validation.type is None:
                continue
            for cell_range in validation.sqref.ranges:
                associated_columns = self._get_column(cell_range)
                for col in associated_columns:
                    if col > len(columns_list):
                        continue
                    predicted_type = self._predict_type(validation)
                    # Will be overriden if conflicting values...
                    validation_dict[columns_list[col - 1]] = predicted_type
        data = self._generate_script(columns_list, validation_dict)
        with open(self.output, "w") as f:
            f.write(data)

    def _get_column(self, cell_range):
        if cell_range.min_row == cell_range.max_row:
            # Might be a mistake, ignore it
            return []
        return set([cell_range.min_col, cell_range.max_col])

    def _predict_type(self, validation):
        if validation.type == "decimal":
            self.used_validators.add("FloatValidator")
            return "FloatValidator({})".format(self._get_numbers_limits(validation))
        if validation.type == "whole":
            self.used_validators.add("IntValidator")
            return "IntValidator({})".format(self._get_numbers_limits(validation))
        if validation.type == "date":
            self.used_validators.add("DateValidator")
            return "DateValidator()"
        if validation.type == "list":
            self.used_validators.add("SetValidator")
            return "SetValidator({})".format(self._get_set_values(validation))
        else:
            return "NoValidator()"

    def _get_numbers_limits(self, validation):
        # Compatibility for "between"
        if validation.operator == "between" or (validation.operator is None and validation.formula1 is not None and validation.formula2 is not None):
            return 'min={}, max={}'.format(validation.formula1, validation.formula2)
        elif validation.operator in ["greaterThan", "greaterThanOrEqual"]:
            return 'min={}'.format(validation.formula1)
        elif validation.operator in ["lessThan", "lessThanOrEqual"]:
            return 'max={}'.format(validation.formula1)
        else:
            return ''

    def _get_set_values(self, validation):
        if "," in validation.formula1:
            value_list = validation.formula1.replace('"', '').split(",")
            value_list = ['"{}"'.format(value) for value in value_list]
            value_string = ', '.join(value_list)
            return 'valid_values=[{}]'.format(value_string)
        # Else it is a cell range : should extract values?
        elif validation.formula1.lstrip("=") in self.names:
            cell_range = self.names[validation.formula1.lstrip("=")]
            value_list = []
            for sheet, coords in cell_range:
                ws = self.wb[sheet]
                for cell_tuple in ws[coords]:
                    for cell in cell_tuple:
                        value_list.append(cell.value)
            value_list = ['"{}"'.format(value) for value in value_list]
            value_string = ', '.join(value_list)
            return 'valid_values=[{}]'.format(value_string)
        else:
            return ""

    def _generate_script(self, columns_list, validation_dict):
        self.used_validators.add("NoValidator")
        validators_list = list(self.used_validators)
        content = ("from checkcel import Checkplate\n"
                   "from checkcel.validators import {}\n"
                   "from collections import OrderedDict\n"
                   "\n"
                   "\n"
                   "class MyTemplate(Checkplate):\n"
                   "    validators = OrderedDict([\n"
                   ).format(", ".join(validators_list))
        for column in columns_list:
            validator = validation_dict.get(column, "NoValidator()")
            content += '        ("{}", {}),\n'.format(column, validator)
        content = content.rstrip(",\n") + "\n"
        content += "    ])\n"
        return content

from email_validator import validate_email, EmailNotValidError
import requests
import re

from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname, column_index_from_string
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName
from urllib.parse import quote_plus
from dateutil import parser

from collections import defaultdict

from checkcel.exceptions import ValidationException, BadValidatorException
from checkcel import logs


class Validator(object):
    """ Generic Validator class """

    def __init__(self, empty_ok=None, ignore_case=None, ignore_space=None):
        self.logger = logs.logger
        self.invalid_dict = defaultdict(set)
        self.fail_count = 0
        self.empty_ok = empty_ok
        self.ignore_case = ignore_case
        self.ignore_space = ignore_space

    @property
    def bad(self):
        raise NotImplementedError

    def validate(self, field, row_number, row):
        """ Validate the given field. Also is given the row context """
        raise NotImplementedError

    def generate(self, column):
        """ Generate an openpyxl Datavalidation entity. Pass the column for custom formulas"""
        raise NotImplementedError

    def describe(self, column_name):
        """ Return a line of text describing allowed values"""
        raise NotImplementedError

    def _set_attributes(self, empty_ok_template, ignore_case_template, ignore_space_template):
        # Override with template value if it was not set (default to None)
        if self.empty_ok is None:
            self.empty_ok = empty_ok_template
        if self.ignore_case is None:
            self.ignore_case = ignore_case_template
        if self.ignore_space is None:
            self.ignore_space = ignore_space_template


class NoValidator(Validator):
    """ No check"""

    def __init__(self, **kwargs):
        super(NoValidator, self).__init__(**kwargs)

    def validate(self, field, row_number, row={}):
        pass

    def generate(self, column):
        return None

    def describe(self, column_name):
        return "{} : Free value".format(column_name)

    @property
    def bad(self):
        return self.invalid_dict


class TextValidator(Validator):
    """ Default validator : will only check if not empty"""

    def __init__(self, **kwargs):
        super(TextValidator, self).__init__(**kwargs)

    def validate(self, field, row_number, row={}):
        if not field and not self.empty_ok:
            raise ValidationException(
                "Field cannot be empty"
            )

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column):
        return None

    def describe(self, column_name):
        return "{} : Free text {}".format(column_name, "(required)" if not self.empty_ok else "")


class CastValidator(Validator):
    """ Validates that a field can be cast to a float """

    def __init__(self, min=None, max=None, **kwargs):
        super(CastValidator, self).__init__(**kwargs)
        self.min = min
        self.max = max

    def validate(self, field, row_number, row={}):
        if self.ignore_space:
            field = field.strip()

        try:
            if field or not self.empty_ok:
                field = self.cast(field)
                if self.min is not None and field < self.min:
                    self.invalid_dict["invalid_set"].add(field)
                    self.invalid_dict["invalid_rows"].add(row_number)
                    raise ValidationException("{} is below min value {}".format(field, self.min))
                if self.max is not None and field > self.max:
                    self.invalid_dict["invalid_set"].add(field)
                    self.invalid_dict["invalid_rows"].add(row_number)
                    raise ValidationException("{} is above max value {}".format(field, self.max))

        except ValueError as e:
            self.invalid_dict["invalid_set"].add(field)
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException(e)

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column):
        params = {"type": self.type}
        if (self.min is not None and self.max is not None):
            params["formula1"] = self.min
            params["formula2"] = self.max
            params["operator"] = "between"
        elif self.min is not None:
            params["formula1"] = self.min
            params["operator"] = "greaterThanOrEqual"
        elif self.max is not None:
            params["formula1"] = self.max
            params["operator"] = "lessThanOrEqual"
        dv = DataValidation(**params)
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        text = "{} : {} number".format(column_name, self.type.capitalize())
        if (self.min is not None and self.max is not None):
            text += " ({} - {})".format(self.min, self.max)
        elif self.min is not None:
            text += " >= {}".format(self.min)
        elif self.max is not None:
            text += " <= {}".format(self.max)

        if not self.empty_ok:
            text += " (required)"
        return text


class FloatValidator(CastValidator):
    """ Validates that a field can be cast to a float """

    def __init__(self, **kwargs):
        super(FloatValidator, self).__init__(**kwargs)
        self.cast = float
        self.type = "decimal"


class IntValidator(CastValidator):
    """ Validates that a field can be cast to an int """

    def __init__(self, **kwargs):
        super(IntValidator, self).__init__(**kwargs)
        self.cast = int
        self.type = "whole"


class SetValidator(Validator):
    """ Validates that a field is in the given set of values """

    def __init__(self, valid_values=set(), **kwargs):
        super(SetValidator, self).__init__(**kwargs)
        self.ordered_values = valid_values
        self.valid_values = set(valid_values)
        if self.empty_ok:
            self.valid_values.add("")

    def validate(self, field, row_number, row={}):
        if self.ignore_case:
            field = field.lower()
        if self.ignore_space:
            field = field.strip()

        if field not in self.valid_values:
            self.invalid_dict["invalid_set"].add(field)
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException(
                "'{}' is invalid".format(field)
            )

    def _set_attributes(self, empty_ok_template, ignore_case_template, ignore_space_template):
        # Override with template value if it was not set (default to None)
        if self.empty_ok is None:
            self.empty_ok = empty_ok_template
        if self.empty_ok:
            self.valid_values.add("")

        if self.ignore_case is None:
            self.ignore_case = ignore_case_template
        if self.ignore_case:
            self.valid_values = set([value.lower() for value in self.valid_values])

        if self.ignore_space is None:
            self.ignore_space = ignore_space_template

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column, column_name="", additional_column=None, additional_worksheet=None):
        # If total length > 256 : need to use cells on another sheet
        if additional_column and additional_worksheet:
            params = {"type": "list"}
            cell = additional_worksheet.cell(column=column_index_from_string(additional_column), row=1, value=column_name)
            cell.font = Font(color="FF0000", bold=True)
            row = 2
            for term in self.ordered_values:
                additional_worksheet.cell(column=column_index_from_string(additional_column), row=row, value=term)
                row += 1
            params["formula1"] = "{}!${}$2:${}${}".format(quote_sheetname(additional_worksheet.title), additional_column, additional_column, row - 1)
        else:
            params = {"type": "list"}
            values = ",".join(self.ordered_values)
            params["formula1"] = '"{}"'.format(values)
        dv = DataValidation(**params)
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        return "{} : (Allowed values : {}) {}".format(column_name, ", ".join(self.ordered_values), "(required)" if not self.empty_ok else "")


class LinkedSetValidator(Validator):
    """ Validates that a field is in the given set of values """

    def __init__(self, linked_column="", valid_values={}, **kwargs):
        super(LinkedSetValidator, self).__init__(**kwargs)
        self.valid_values = valid_values
        self.linked_column = linked_column
        self.column_check = False

    def _precheck_unique_with(self, row):
        if self.linked_column not in row.keys():
            raise BadValidatorException("Linked column {} is not in file columns".format(self.linked_column))
        self.column_check = True

    def validate(self, field, row_number, row):
        if self.ignore_case:
            field = field.lower()
        if self.ignore_space:
            field = field.strip()

        if not self.column_check:
            self._precheck_unique_with(row)
        if field == "" and self.empty_ok:
            return
        related_column_value = row[self.linked_column]
        if not related_column_value:
            self.invalid_dict["invalid_rows"].add(row_number)
            self.invalid_dict["invalid_set"].add("Invalid linked column value: ''")
            raise ValidationException("Linked column {} is empty".format(self.linked_column))
        if related_column_value not in self.valid_values.keys():
            self.invalid_dict["invalid_set"].add("Invalid linked column value: {}".format(related_column_value))
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException("Linked column {} value {} is not in valid values".format(self.linked_column, related_column_value))
        if field not in self.valid_values[related_column_value]:
            self.invalid_dict["invalid_set"].add(field)
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException("Value {} is not in allowed values".format(field))

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column, set_columns, column_name, additional_column, additional_worksheet, workbook):
        if self.linked_column not in set_columns:
            # TODO raise warning
            return None
        params = {"type": "list"}
        additional_worksheet.cell(column=column_index_from_string(additional_column), row=1, value=column_name).font = Font(color="FF0000", bold=True)
        row = 2
        row_dict = {}
        for key, value in self.valid_values.items():
            additional_worksheet.cell(column=column_index_from_string(additional_column), row=row, value=key).font = Font(color="FF0000", italic=True)
            row += 1
            row_dict[key] = {'min': row}
            for val in value:
                additional_worksheet.cell(column=column_index_from_string(additional_column), row=row, value=val)
                row += 1
            row_dict[key]['max'] = row - 1
        for key, values in row_dict.items():
            new_range = DefinedName(key, attr_text='{}!${}${}:${}${}'.format(quote_sheetname(additional_worksheet.title), additional_column, values['min'], additional_column, values['max']))
            workbook.defined_names.append(new_range)
        params["formula1"] = "INDIRECT(${}2)".format(set_columns[self.linked_column])
        dv = DataValidation(**params)
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        return "{} : Linked values to column {} {}".format(column_name, self.linked_column, "(required)" if not self.empty_ok else "")


class DateValidator(Validator):
    """ Validates that a field is a Date """

    def __init__(self, day_first=True, **kwargs):
        super(DateValidator, self).__init__(**kwargs)
        self.day_first = day_first

    def validate(self, field, row_number, row={}):
        if self.ignore_space:
            field = field.strip()

        try:
            if field or not self.empty_ok:
                # Pandas auto convert fields into dates (ignoring the parse_dates=False)
                field = str(field)
                parser.parse(field, dayfirst=self.day_first).date()

        except parser.ParserError as e:
            self.invalid_dict["invalid_set"].add(field)
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException(e)

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column, additional_column=None, additional_worksheet=None):
        # GreaterThanOrEqual for validity with ODS.
        dv = DataValidation(type="date", formula1='01/01/1900', operator='greaterThanOrEqual')
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        return "{} : Date {}".format(column_name, "(required)" if not self.empty_ok else "")


class TimeValidator(Validator):
    """ Validates that a field is a Time """

    def __init__(self, **kwargs):
        super(TimeValidator, self).__init__(**kwargs)

    def validate(self, field, row_number, row={}):
        if self.ignore_space:
            field = field.strip()
        try:
            if field or not self.empty_ok:
                # Pandas auto convert fields into dates (ignoring the parse_dates=False)
                field = str(field)
                parser.parse(field).time()

        except parser.ParserError as e:
            self.invalid_dict["invalid_set"].add(field)
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException(e)

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column, additional_column=None, additional_worksheet=None):
        # GreaterThanOrEqual for validity with ODS.
        dv = DataValidation(type="time")
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        return "{} : Time {}".format(column_name, "(required)" if not self.empty_ok else "")


class EmailValidator(Validator):
    """ Validates that a field is in the given set """

    def __init__(self, **kwargs):
        super(EmailValidator, self).__init__(**kwargs)

    def validate(self, field, row_number, row={}):
        if self.ignore_space:
            field = field.strip()
        if field or not self.empty_ok:
            try:
                validate_email(field)
            except EmailNotValidError as e:
                self.invalid_dict["invalid_set"].add(field)
                self.invalid_dict["invalid_rows"].add(row_number)
                raise ValidationException(e)

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column, ontology_column=None):
        params = {"type": "custom"}
        params["formula1"] = '=ISNUMBER(MATCH("*@*.?*",{}2,0))'.format(column)
        dv = DataValidation(**params)
        dv.error = 'Value must be an email'
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        return "{} : Email {}".format(column_name, "(required)" if not self.empty_ok else "")


class OntologyValidator(Validator):
    """ Validates that a field is in the given set """

    def __init__(self, ontology, root_term="", **kwargs):
        super(OntologyValidator, self).__init__(**kwargs)
        self.validated_terms = set()
        self.ontology = ontology
        self.root_term = root_term
        self.root_term_iri = ""

        is_ontology, self.root_term_iri = self._validate_ontology()
        if not is_ontology:
            raise BadValidatorException("'{}' is not a valid ontology".format(self.ontology))
        if self.root_term and not self.root_term_iri:
            raise BadValidatorException("'{}' is not a valid root term for ontology {}".format(self.root_term, self.ontology))

    def validate(self, field, row_number, row={}):
        if self.ignore_space:
            field = field.strip()

        if field == "" and self.empty_ok:
            return

        if field in self.invalid_dict["invalid_set"]:
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException("{} is not an ontological term".format(field))

        if field not in self.validated_terms:
            ontological_term = self._validate_ontological_term(field)
            if not ontological_term:
                self.invalid_dict["invalid_set"].add(field)
                self.invalid_dict["invalid_rows"].add(row_number)
                raise ValidationException("{} is not an ontological term".format(field))
            self.validated_terms.add(field)

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column, additional_column, additional_worksheet):
        terms = self._get_ontological_terms()
        cell = additional_worksheet.cell(column=column_index_from_string(additional_column), row=1, value=self.ontology)
        cell.font = Font(color="FF0000", bold=True)
        row = 2
        for term in terms:
            additional_worksheet.cell(column=column_index_from_string(additional_column), row=row, value=term)
            row += 1

        params = {"type": "list"}
        params["formula1"] = "{}!${}$2:${}${}".format(quote_sheetname(additional_worksheet.title), additional_column, additional_column, row - 1)
        dv = DataValidation(**params)
        dv.error = 'Value must be an ontological term'
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        text = "{} : Ontological term from {} ontology.".format(column_name, self.ontology)
        if self.root_term:
            text += " Root term is : {}".format(self.root_term)
        if not self.empty_ok:
            text += " (required)"
        return text

    def _validate_ontological_term(self, term, return_uri=False):
        base_path = "http://www.ebi.ac.uk/ols/api/search"
        body = {
            "q": term,
            "ontology": self.ontology.lower(),
            "type": "class",
            "exact": True,
            "queryFields": ["label", "synonym"]
        }
        if self.root_term_iri:
            body["childrenOf"] = self.root_term_iri
        r = requests.get(base_path, params=body)
        res = r.json()
        if not res["response"]["numFound"] == 1:
            return False
        if return_uri:
            return res["response"]["docs"][0]["iri"]
        return True

    def _get_ontological_terms(self):
        size = 100
        terms = set()
        if self.root_term_iri:
            url = "http://www.ebi.ac.uk/ols/api/ontologies/{}/terms/{}/descendants?size={}".format(self.ontology, quote_plus(quote_plus(self.root_term_iri)), size)
        else:
            url = "http://www.ebi.ac.uk/ols/api/ontologies/{}/terms?size={}".format(self.ontology, size)

        r = requests.get(url)
        res = r.json()
        for term in res["_embedded"]["terms"]:
            terms.add(term["label"])
        while "next" in res["_links"]:
            url = res["_links"]["next"]["href"]
            r = requests.get(url)
            res = r.json()
            for term in res["_embedded"]["terms"]:
                terms.add(term["label"])

        return terms

    def _validate_ontology(self):
        root_term_iri = ""
        if not self.ontology:
            return False
        base_path = "http://www.ebi.ac.uk/ols/api"
        sub_path = "/ontologies/{}".format(self.ontology.lower())
        r = requests.get(base_path + sub_path)
        if not r.status_code == 200:
            return False, root_term_iri
        if self.root_term:
            root_term_iri = self._validate_ontological_term(self.root_term, return_uri=True)
        return True, root_term_iri


class UniqueValidator(Validator):
    """ Validates that a field is unique within the file """

    def __init__(self, unique_with=[], **kwargs):
        super(UniqueValidator, self).__init__(**kwargs)
        self.unique_values = set()
        self.unique_with = unique_with
        self.unique_check = False

    def _precheck_unique_with(self, row):
        extra = set(self.unique_with) - set(row.keys())
        if extra:
            raise BadValidatorException(extra)
        self.unique_check = True

    def validate(self, field, row_number, row={}):
        if self.ignore_space:
            field = field.strip()

        if not field and not self.empty_ok:
            raise ValidationException(
                "Field cannot be empty"
            )

        if self.unique_with and not self.unique_check:
            self._precheck_unique_with(row)

        key = tuple([field] + [row[k] for k in self.unique_with])
        if key not in self.unique_values:
            self.unique_values.add(key)
        else:
            self.invalid_dict["invalid_set"].add(field)
            self.invalid_dict["invalid_rows"].add(row_number)
            if self.unique_with:
                raise ValidationException(
                    "'{}' is already in the column (unique with: {})".format(
                        field, key[1:]
                    )
                )
            else:
                raise ValidationException("'{}' is already in the column".format(field))

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column, column_dict):
        if self.unique_with and not all([val in column_dict for val in self.unique_with]):
            raise BadValidatorException("Using unique_with, but the related column was not defined before")

        params = {"type": "custom"}
        internal_value = "${0}:${0},{0}2".format(column)
        if self.unique_with:
            for col in self.unique_with:
                internal_value += ",${0}:${0},{0}2".format(column_dict[col])
        params["formula1"] = '=COUNTIF({})<2'.format(internal_value)
        dv = DataValidation(**params)
        dv.error = 'Value must be unique'
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        text = "{} : Unique value".format(column_name)
        if self.unique_with:
            text += " Must be unique with column(s) {}".format(", ".join(self.unique_with))
        if not self.empty_ok:
            text += " (required)"
        return text


class VocabulaireOuvertValidator(Validator):
    """ Validates that a term is part of the INRAE thesaurus """

    def __init__(self, root_term="", lang="en", labellang="en", vocab="thesaurus-inrae", **kwargs):
        super(VocabulaireOuvertValidator, self).__init__(**kwargs)
        self.validated_terms = set()
        self.root_term = root_term
        self.root_term_iri = ""
        self.lang = lang
        self.labellang = labellang if labellang else self.lang
        self.vocab = vocab

        if self.vocab:
            # Check if vocab exist here
            if not self._validate_vo_vocab():
                raise BadValidatorException("'{}' is not a valid vocabulary".format(self.ontology))

        if self.root_term:
            exists, self.root_term_iri = self._validate_vo_term(self.root_term, return_uri=True)
            if not exists:
                raise BadValidatorException("'{}' is not a valid root term. Make sure it is a concept, and not a microthesaurus or group".format(self.root_term))

    def validate(self, field, row_number, row={}):
        if self.ignore_space:
            field = field.strip()

        if field == "" and self.empty_ok:
            return

        if field in self.invalid_dict["invalid_set"]:
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException("{} is not an ontological term".format(field))

        if field not in self.validated_terms:
            ontological_term, _ = self._validate_vo_term(field)
            if not ontological_term:
                self.invalid_dict["invalid_set"].add(field)
                self.invalid_dict["invalid_rows"].add(row_number)
                raise ValidationException("{} is not an ontological term".format(field))
            self.validated_terms.add(field)

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column, additional_column, additional_worksheet):
        # No point in loading 15000 terms
        # No easy way to do it anyway
        if not self.root_term_iri:
            self.logger.warning(
                "Warning: no root term used. No validation will be generated"
            )
            return None

        terms = self._get_vo_terms()

        if not terms:
            self.logger.warning(
                "Warning: 0 descendants found for root term {}. It might not be a concept".format(self.root_term)
            )
            return None

        cell = additional_worksheet.cell(column=column_index_from_string(additional_column), row=1, value=self.vocab)
        cell.font = Font(color="FF0000", bold=True)
        row = 2
        for term in terms:
            additional_worksheet.cell(column=column_index_from_string(additional_column), row=row, value=term)
            row += 1

        params = {"type": "list"}
        params["formula1"] = "{}!${}$2:${}${}".format(quote_sheetname(additional_worksheet.title), additional_column, additional_column, row - 1)
        dv = DataValidation(**params)
        dv.error = 'Value must be from Vocabulaires ouverts'
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        text = "{} : Ontological term from Vocabulaires ouverts.".format(column_name)
        if self.root_term:
            text += " Root term is : {}".format(self.root_term)
        if not self.empty_ok:
            text += " (required)"
        return text

    def _validate_vo_term(self, field, return_uri=False):
        params = {"query": field, "unique": True, "type": "skos:Concept"}
        if self.root_term_iri:
            params["parent"] = self.root_term_iri
        if self.lang:
            params["lang"] = self.lang
        if self.labellang:
            params["labellang"] = self.labellang
        if self.vocab:
            params["vocab"] = self.vocab

        url = "https://consultation.vocabulaires-ouverts.inrae.fr/rest/v1/search"

        r = requests.get(url, params=params)
        res = r.json()
        # Might be a better way. Check prefLabel?
        if not len(res["results"]) == 1:
            return False, ""

        if return_uri:
            return True, res["results"][0]['uri']
        return True, ""

    def _get_vo_terms(self):
        url = "https://consultation.vocabulaires-ouverts.inrae.fr/rest/v1/{}/narrowerTransitive".format(self.vocab)
        params = {"uri": self.root_term_iri}

        if self.lang:
            params['lang'] = self.lang

        r = requests.get(url, params=params)
        res = r.json()

        return sorted([term['prefLabel'] for term in res['narrowerTransitive'].values() if term.get('prefLabel')])

    def _validate_vo_vocab(self):
        url = "https://consultation.vocabulaires-ouverts.inrae.fr/rest/v1/" + self.vocab
        r = requests.get(url)

        if not r.status_code == 200:
            return False

        res = r.json()
        if not res['type'] and res['type'][0]['prefLabel'] == "Thesaurus":
            return False

        return True


class RegexValidator(Validator):
    """ Validates that a term match a regex"""

    def __init__(self, regex, excel_formula="", **kwargs):
        super(RegexValidator, self).__init__(**kwargs)
        self.regex = regex
        self.excel_formula = excel_formula
        try:
            re.compile(regex)
        except re.error:
            raise BadValidatorException("'{}' is not a valid regular expression".format(self.regex))

    def validate(self, field, row_number, row={}):
        if self.ignore_space:
            field = field.strip()

        if field == "" and self.empty_ok:
            return

        matches = re.findall(self.regex, field)
        if not len(matches) == 1:
            self.invalid_dict["invalid_set"].add(field)
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException("{} does not match regex {}".format(field, self.regex))

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column):
        # Difficult to use regex in Excel without a VBA macro
        if not self.excel_formula:
            self.logger.warning(
                "Warning: RegexValidator does not generate a validated column"
            )
            return None

        params = {"type": "custom"}
        params["formula1"] = self.excel_formula.replace("{CNAME}", column)
        dv = DataValidation(**params)
        dv.error = 'Value must match validation'
        dv.add("{}2:{}1048576".format(column, column))
        return dv

    def describe(self, column_name):
        text = "{} : Term matching the regex {}.".format(column_name, self.regex)
        if not self.empty_ok:
            text += " (required)"
        return text


class GPSValidator(Validator):
    """ Validates that a term match a regex"""

    def __init__(self, format="DD", only_long=False, only_lat=False, **kwargs):
        super(GPSValidator, self).__init__(**kwargs)
        self.format = format

        if format not in ['DD', 'DMS']:
            raise BadValidatorException("Error: Format must be in 'DD' or 'DMS' format")

        if only_long and only_lat:
            raise BadValidatorException("Error: cannot set both only_long and only_lat")

        self.only_long = only_long
        self.only_lat = only_lat

    def validate(self, field, row_number, row={}):
        if self.ignore_space:
            field = field.strip()

        if field == "" and self.empty_ok:
            return

        if self.format == "DD":
            regex_lat = r"[-+]?((90(\.0+)?)|([1-8]?\d(\.\d+)?))[NSns]?"
            regex_long = r"[-+]?((180(\.0+)?)|(((1[0-7]\d)|([1-9]?\d))(\.\d+)?))[wWeE]?"

        else:
            regex_lat = r"((([1-8]?\d)(°\s?|\s)([1-5]?\d|60)('\s?|\s)?([1-5]?\d(\.\d+)?|60)(\"\s?|\s)?)|(90(°\s?|\s)0('\s?|\s)0(\"\s?|\s)?))[NSns]?"
            regex_long = r"((((1[0-7][0-9])|([0-9]{1,2}))(°\s?|\s)([1-5]?\d|60)('\s?|\s)([1-5]?\d(\.\d+)?|60)(\"\s?|\s)?)|(180(°\s?|\s)0('\s?|\s)0(\"\s?|\s)?))[EWew]?"

        if self.only_long:
            regex = r"^{}$".format(regex_long)
        elif self.only_lat:
            regex = r"^{}$".format(regex_lat)
        else:
            regex = r"^{}[,\s]?\s?{}$".format(regex_lat, regex_long)

        matches = re.findall(regex, field)
        if not len(matches) == 1:
            self.invalid_dict["invalid_set"].add(field)
            self.invalid_dict["invalid_rows"].add(row_number)
            raise ValidationException("{} is not a valid GPS coordinate")

    @property
    def bad(self):
        return self.invalid_dict

    def generate(self, column):
        # Difficult to use regex in Excel without a VBA macro
        self.logger.warning(
            "Warning: GPSValidator does not generate a validated column"
        )
        return None

    def describe(self, column_name):
        text = "{} : GPS coordinate".format(column_name)
        if not self.empty_ok:
            text += " (required)"
        return text
